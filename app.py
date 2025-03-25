from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
import os
import re
import shutil
import json
import cv2
import hashlib
import requests
import threading
from openpyxl import Workbook, load_workbook
import pandas as pd
import pytesseract as ts
from selenium.common.exceptions import NoSuchElementException

# ----- CONFIGURAÇÃO INICIAL -----

# Caminhos definidos diretamente no código
chromedriver_path = os.path.join(os.getenv("ProgramFiles"), "Google", "chromedriver-win64", "chromedriver.exe")
chrome_profile_path = os.path.join(os.getenv("LOCALAPPDATA"), "Google", "Chrome", "User Data", "Default")
base_folder_path = os.path.join(os.getenv("USERPROFILE"), "Downloads", "Comprovantes")
downloads_folder_path = os.path.join(os.getenv("USERPROFILE"), "Downloads")
tesseract_path = shutil.which("tesseract") or os.path.join(os.getenv("ProgramFiles"), "Tesseract-OCR", "tesseract.exe")

# Configurar o Tesseract
ts.pytesseract.tesseract_cmd = tesseract_path
# Definir o idioma para o Tesseract (Português)
tessdata_path = "C:/Program Files/Tesseract-OCR/tessdata"
# Confirmar que o caminho do executável está correto
ts.pytesseract.tesseract_cmd = tesseract_path
# Definir o caminho correto da tessdata
os.environ['TESSDATA_PREFIX'] = tessdata_path


imagens_processadas = set()
imagens_baixadas = []
programa_ativo = True


# Função para limpar a pasta de registros antigos
#def limpar_pasta(pasta):
   # if os.path.exists(pasta):
       # for arquivo in os.listdir(pasta):
           # caminho_arquivo = os.path.join(pasta, arquivo)
            #try:
              #  if os.path.isfile(caminho_arquivo):
               #     os.remove(caminho_arquivo)
           # except Exception as e:
               # print(f"⚠️ Erro ao remover {arquivo}: {e}")


# Limpa a pasta antes de iniciar se necessário
#limpar_pasta("Comprovantes")
#print("🧹 Todos os arquivos da pasta 'Comprovantes' foram removidos!")

# Configuração do Selenium
chrome_options = Options()
chrome_options.add_argument(f"--user-data-dir={chrome_profile_path}")
service = Service(chromedriver_path)

# Iniciar o navegador
navegador = webdriver.Chrome(service=service, options=chrome_options)
navegador.get("https://web.whatsapp.com/")

# Aguardar login do usuário
input("📲 Escaneie o QR Code e pressione Enter para continuar...")
print("Digite 'e' para encerrar o programa e limpar as imagens")

# Criar diretório
pasta_base = "Comprovantes"
os.makedirs(pasta_base, exist_ok=True)

# Definição dos arquivos Excel
xlsx_funcionario = os.path.join(pasta_base, "comprovantes_funcionario.xlsx")
xlsx_motoboy = os.path.join(pasta_base, "comprovantes_motoboy.xlsx")

# Função para criar arquivos Excel caso não existam
def criar_arquivo_excel(arquivo):
    if not os.path.exists(arquivo):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Horário", "Valor",
                  "Destinatário", "Categoria"])
        wb.save(arquivo)
        print(f"📊 Arquivo criado: {arquivo}")


# Criar arquivos Excel
criar_arquivo_excel(xlsx_funcionario)
criar_arquivo_excel(xlsx_motoboy)

# Função para ajustar a largura das colunas no Excel
def ajustar_largura_colunas(arquivo):
    wb = load_workbook(arquivo)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(arquivo)

# Função para carregar mensagens já salvas no Excel
def carregar_mensagens_anteriores(arquivo):
    wb = load_workbook(arquivo)
    ws = wb.active
    mensagens_existentes = set()

    # Pular a primeira linha de cabeçalho
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome, horario, valor, destinatario, mensagem = row
        identificador = f"{nome}-{horario}-{valor}-{mensagem}"
        mensagens_existentes.add(identificador)

    return mensagens_existentes

# Conjunto para armazenar mensagens já processadas
mensagens_processadas = set()

# Função para classificar categoria
def classificar_categoria(mensagem):
    if "motoboy" in mensagem.lower():
        return "Motoboy"
    elif "funcionario" or "funcionário" in mensagem.lower():
        return "Funcionário"
    return "Outros"

# Função para extrair valor do Pix
def extrair_valor(mensagem):
    padrao_valor = r'R\$\s*\d{1,3}(?:\.\d{3})*(?:,\d{2})?'
    valores = re.findall(padrao_valor, mensagem)
    return valores[0].replace(" ", "") if valores else "Não encontrado"

# Função para remover erros de codificação
def corrigir_acentuacao(texto):
    return texto.encode('utf-8').decode('utf-8-sig')


def gerar_nome_arquivo_hash(identificador):
    return hashlib.md5(identificador.encode()).hexdigest() + ".jpeg"


def baixar_imagem(imagem_elemento, nome_arquivo):
    """Faz o download da imagem, usando JavaScript se for URL blob ou requests."""
    try:
        imagem_url = imagem_elemento.get_attribute("src")
        if imagem_url.startswith("blob:"):
            script = f"""
            var xhr = new XMLHttpRequest();
            xhr.open('GET', '{imagem_url}', true);
            xhr.responseType = 'blob';
            xhr.onload = function() {{
                var blob = xhr.response;
                var reader = new FileReader();
                reader.onloadend = function() {{
                    var base64data = reader.result;
                    var link = document.createElement('a');
                    link.href = base64data;
                    link.download = '{nome_arquivo}';
                    link.click();
                }}; 
                reader.readAsDataURL(blob);
            }}; 
            xhr.send();
            """
            navegador.execute_script(script)
            imagens_baixadas.append(nome_arquivo)
            print(f"✅ Imagem salva como: {nome_arquivo}")
        else:
            resposta = requests.get(imagem_url)
            if resposta.status_code == 200:
                with open(nome_arquivo, 'wb') as f:
                    f.write(resposta.content)
                imagens_baixadas.append(nome_arquivo)
                print(f"✅ Imagem salva como: {nome_arquivo}")
            else:
                print(
                    f"⚠️ Erro ao baixar a imagem: Status {resposta.status_code}")
    except Exception as e:
        print(f"⚠️ Erro ao baixar a imagem: {e}")


def analisar_imagem(nome_imagem, pasta_downloads=downloads_folder_path, tentativas_max=10, intervalo_espera=1):
    """Tenta localizar e analisar a imagem na pasta de downloads, com múltiplas tentativas."""
    for tentativa in range(tentativas_max):
        try:
            caminho_imagem = os.path.join(pasta_downloads, nome_imagem)
            arquivos_na_pasta = os.listdir(pasta_downloads)

            arquivo_encontrado = None
            for arquivo in arquivos_na_pasta:
                if (arquivo.lower().replace('ç', 'c') == nome_imagem.lower().replace('ç', 'c')): 
                    arquivo_encontrado = os.path.join(pasta_downloads, arquivo)
                    break

            if not arquivo_encontrado:
                print(f"⚠️ Imagem não encontrada na tentativa {tentativa + 1}. Aguardando...")
                time.sleep(intervalo_espera)
                continue

            caminho_imagem = arquivo_encontrado
            print(f"Arquivo encontrado: {caminho_imagem}")

            if os.path.getsize(caminho_imagem) == 0:
                print(f"⚠️ Arquivo vazio na tentativa {tentativa + 1}. Aguardando...")
                time.sleep(intervalo_espera)
                continue

            img = cv2.imread(caminho_imagem)
            if img is None:
                print(f"⚠️ Erro ao carregar a imagem na tentativa {tentativa + 1}. Aguardando...")
                time.sleep(intervalo_espera)
                continue

            text_img = ts.image_to_string(img, lang='por')
            valor = re.findall(r'R\$\s*\d+[\.,]?\d*', text_img)
            destinatario = re.findall(r'(?i)(?:para|nome do favorecido)\s*:?\s*([^\n]+)', text_img)

            if valor:
                valor = valor[0].strip()
            else:
                valor = "Valor não encontrado"

            if destinatario:
                ditemp = destinatario[0].strip().replace('\n', ' ')
                destinatario = ' '.join(ditemp.split())
            else:
                destinatario = "Destinatário não encontrado"
        ########    print(f"Valor/Destinario: | {valor} | {destinatario}|")
            return valor, destinatario

        except Exception as e:
            print(f"Erro na tentativa {tentativa + 1}: {e}")
            time.sleep(intervalo_espera)

    print(f"⚠️ Falha ao processar a imagem {nome_imagem} após {tentativas_max} tentativas.")
    return None, None

# Função para extrair mensagens

from openpyxl import load_workbook

def obter_ultimo_horario(arquivo_excel):
    """ Retorna o último horário armazenado no Excel para evitar duplicação. """
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active

        horarios = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Pulando cabeçalho
            if row[1]:  # Coluna de horário (índice 1)
                horarios.append(row[1])

        wb.close()
        if horarios:
            return max(horarios)  # Retorna o maior horário encontrado
        return None
    except Exception as e:
        print(f"⚠️ Erro ao obter o último horário do arquivo {arquivo_excel}: {e}")
        return None



def extrair_mensagens():
    novas_mensagens = []

    bolhas = navegador.find_elements(By.XPATH, '//div[contains(@class, "message-in") or contains(@class, "message-out")]')

        # Obter o último horário registrado para cada categoria
    ultimo_horario_funcionario = obter_ultimo_horario(xlsx_funcionario)
    ultimo_horario_motoboy = obter_ultimo_horario(xlsx_motoboy)


    for bolha in bolhas:
        try:
            nome = "Você"
            horario = "Desconhecido"

            if "message-in" in bolha.get_attribute("class"):
                nome_elemento = bolha.find_element(By.XPATH, './/div[contains(@class, "copyable-text")]')
                nome = nome_elemento.get_attribute("data-pre-plain-text")

                if "[" in nome and "]" in nome:
                    horario = nome.split("]")[0].replace("[", "").strip()
                    nome = nome.split("] ")[-1].strip()

            texto_elemento = bolha.find_elements(By.XPATH, './/span[contains(@class, "selectable-text")]')
            texto = " ".join([t.text for t in texto_elemento]).strip()

            if texto.startswith("Transferência realizada"):
                categoria = classificar_categoria(texto)
                valor = extrair_valor(texto)
                identificador = f"{nome}-{horario}-{valor}-{texto}"

               # Verificar se a mensagem é mais recente que o último horário salvo

                if categoria == "Funcionário" and ultimo_horario_funcionario and horario <= ultimo_horario_funcionario:
                    continue
                if categoria == "Motoboy" and ultimo_horario_motoboy and horario <= ultimo_horario_motoboy:
                    continue

                try:
                    imagem_elemento = bolha.find_element(By.XPATH, './/img[contains(@src, "blob:") or contains(@class, "media")]')
                except NoSuchElementException:
                    imagem_elemento = None

                if imagem_elemento:
                    nome_arquivo = gerar_nome_arquivo_hash(identificador)
 

                if nome_arquivo not in mensagens_processadas:
                    baixar_imagem(imagem_elemento, nome_arquivo)
                    valor,destinatario= analisar_imagem(nome_arquivo)
                    mensagens_processadas.add(nome_arquivo)
                    novas_mensagens.append((nome, horario, categoria, valor, destinatario, texto))

        except Exception as e:
            print(f"⚠️ Erro ao extrair mensagem: {e}")

    return novas_mensagens


def limpar_imagens_baixadas(pasta, imagens_ids):
    for nome_arquivo in imagens_ids:
        caminho_arquivo = os.path.join(pasta, nome_arquivo)
        try:
            if os.path.isfile(caminho_arquivo):
                os.remove(caminho_arquivo)
                print(f"🗑️ Imagem removida: {nome_arquivo}")
        except Exception as e:
            print(f"⚠️ Erro ao remover {nome_arquivo}: {e}")

def monitorar_entrada():
    global programa_ativo
    while programa_ativo:
        comando = input().strip().lower()
        if comando == 'e':
            print("🛑 Comando 'e' recebido. Encerrando o programa...")
            programa_ativo = False


thread_monitoramento = threading.Thread(target=monitorar_entrada, daemon=True)
thread_monitoramento.start()

# Monitoramento contínuo
while programa_ativo:
    mensagens = extrair_mensagens()

    for nome, horario, categoria, valor, destinatario, mensagem in mensagens:
        if categoria == "Funcionário":
            arquivo_excel = xlsx_funcionario
        elif categoria == "Motoboy":
            arquivo_excel = xlsx_motoboy
        else:
            continue

        # Carregar mensagens já salvas para evitar duplicação
        mensagens_existentes = carregar_mensagens_anteriores(arquivo_excel)

        identificador = f"{nome}-{horario}-{valor}-{mensagem}"
        if identificador not in mensagens_existentes:
            print(f"💾 Salvando no arquivo: {arquivo_excel}")

            try:
                wb = load_workbook(arquivo_excel)
                ws = wb.active
                ws.append([nome, horario, valor, destinatario, categoria])
                wb.save(arquivo_excel)
                print(f"✅ Mensagem salva: {mensagem}")

                # Ajustar a largura das colunas após salvar
                ajustar_largura_colunas(arquivo_excel)

            except Exception as e:
                print(f"⚠️ Erro ao salvar no Excel: {e}")
        else:
            print("🛑 Mensagem já salva, ignorando.")

    time.sleep(5)

limpar_imagens_baixadas(downloads_folder_path, imagens_baixadas)
